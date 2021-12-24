import openpyxl
from random import randint
import random

pedidos = openpyxl.load_workbook(r'C:\Users\vynic\PycharmProjects\UdemyModulos\Ler_planilhas_excel\pedidos.xlsx')


planilha = pedidos['Página1']

for linha in range(5,16):
    num_pedido = linha - 1
    planilha.cell(linha, 1).value = num_pedido

for linha in range(5,16):
    planilha.cell(linha, 2).value = randint(1000,1300)

for linha in range(5,16):
    #round = Arredondar o numero para X casas decimais
    random_price = round(random.uniform(50,200), 2)
    planilha.cell(linha, 3).value = random_price





pedidos.save('planilha_do_zero.xlsx')