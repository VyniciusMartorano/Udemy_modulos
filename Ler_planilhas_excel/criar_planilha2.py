import openpyxl
from random import randint, uniform

planilha = openpyxl.Workbook()
planilha.create_sheet('planilha1',0)


planilha1 = planilha['planilha1']


for linha in range(1,11):
    planilha1.cell(linha, 1).value = linha
    num_pedido = 1200 + linha
    planilha1.cell(linha, 2).value = num_pedido
    planilha1.cell(linha, 3).value = f'R${round(uniform(50,300), 2)}'



planilha.save('planilha2.xlsx')

